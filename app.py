import streamlit as st
import io, zipfile, re
from datetime import datetime
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# Leer datos_entrada.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def leer_datos(f):
    wb = load_workbook(f, data_only=True)
    ws = wb["DATOS_PROYECTO"]
    def v(ref):
        val = ws[ref].value
        if val is None: return ""
        if isinstance(val, datetime): return val.strftime("%d/%m/%Y")
        return str(val).strip()
    return {
        "num_rev":         v("B2"),   # Número de revisión
        # Portada
        "titulo_portada":  v("B5"),   # Título portada (CuadroTexto 4 run 3)
        "tramo_portada":   v("B6"),   # Tramo portada  (CuadroTexto 4 run 2)
        # Control de firmas
        "nombre_realizado":v("B9"),
        "nombre_revisado": v("B10"),
        "nombre_aprobado": v("B11"),
        "fecha_firmas":    v("B12"),
        # Registro de cambios
        "autor_cambio":    v("B15"),
        "seccion_afectada":v("B16"),
        "desc_cambio":     v("B17"),
        # Cabecera HRC
        "cod_rev_general": v("B20"),  # F7  + header No.Doc.
        "clave_ref":       v("B21"),  # F9
        "titulo_hrc":      v("B22"),  # F10 — distinto al de portada
        "tramo_hrc":       v("B23"),  # C7  — México-Querétaro-Irapuato...
        "cod_documento":   v("B24"),  # F11
        "originador":      v("B25"),  # F12
        "revisor_ppal":    v("B26"),  # J11
        "entidad_rev":     v("B27"),  # J12
        "fecha_ciclo":     v("B30"),  # N11/O11/P11 según ciclo
    }

# ─────────────────────────────────────────────────────────────────────────────
# Helpers — todo string replace, nunca ET.tostring()
# ─────────────────────────────────────────────────────────────────────────────
def get_sheet_map(files):
    wb_xml = files["xl/workbook.xml"].decode("utf-8")
    sheets = re.findall(r'<sheet\s+name="([^"]+)"[^/]*/>', wb_xml)
    return {name: i+1 for i, name in enumerate(sheets)}

def get_drawing_path(files, sheet_num):
    rpath = f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels"
    if rpath not in files: return None
    rels  = files[rpath].decode("utf-8")
    match = re.search(r'Type="[^"]*relationships/drawing"[^>]*Target="([^"]+)"', rels)
    return match.group(1).replace("../", "xl/") if match else None

# ── PORTADA: CuadroTexto 4 — run 2 (tramo) y run 3 (título) ─────────────────
def actualizar_portada(drawing_bytes, titulo, tramo):
    xml   = drawing_bytes.decode("utf-8")
    match = re.search(r'name="CuadroTexto 4".*?</xdr:sp>', xml, re.DOTALL)
    if not match:
        return drawing_bytes, False, "CuadroTexto 4 no encontrado"
    shape  = match.group(0)
    textos = re.findall(r'<a:t>([^<]*)</a:t>', shape)
    if len(textos) < 4:
        return drawing_bytes, False, f"solo {len(textos)} runs encontrados"

    shape_nuevo = shape
    # run 2 (índice 2) = tramo
    if tramo:
        t_esc = tramo.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
        shape_nuevo = shape_nuevo.replace(
            f"<a:t>{textos[2]}</a:t>", f"<a:t>{t_esc}</a:t>", 1)
    # run 3 (índice 3) = título
    if titulo:
        t_esc = titulo.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
        # Reemplazar en shape_nuevo (ya con tramo cambiado)
        textos_nuevos = re.findall(r'<a:t>([^<]*)</a:t>', shape_nuevo)
        shape_nuevo = shape_nuevo.replace(
            f"<a:t>{textos_nuevos[3]}</a:t>", f"<a:t>{t_esc}</a:t>", 1)

    xml_nuevo = xml[:match.start()] + shape_nuevo + xml[match.end():]
    return xml_nuevo.encode("utf-8"), True, f"título='{titulo}' tramo='{tramo}'"

# ── HEADER DE PÁGINA ──────────────────────────────────────────────────────────
def actualizar_odd_header(sheet_bytes, cod_rev_general, num_rev):
    xml_str  = sheet_bytes.decode("utf-8")
    cod_base = re.sub(r"-S\d{2}$", "", cod_rev_general).strip()
    rev_str  = "S" + num_rev.zfill(2)
    nueva, n = re.subn(
        r"(No\.[ ]Doc\.[ ])[^\r\n]+(\r\n)(Rev\.[ ])S\d{2}",
        lambda m: f"{m.group(1)}{cod_base} {m.group(2)}{m.group(3)}{rev_str}",
        xml_str
    )
    return nueva.encode("utf-8"), n > 0

# ── SHARED STRINGS ────────────────────────────────────────────────────────────
def get_shared_strings(files):
    ss_xml = files["xl/sharedStrings.xml"].decode("utf-8")
    items  = re.findall(r'<si>(.*?)</si>', ss_xml, re.DOTALL)
    return [''.join(re.findall(r'<t[^>]*>([^<]*)</t>', it)) for it in items]

def rebuild_shared_strings(shared):
    items = []
    for s in shared:
        s_esc = (s.replace("&","&amp;").replace("<","&lt;")
                  .replace(">","&gt;").replace('"',"&quot;"))
        space = ' xml:space="preserve"' if s != s.strip() or '\n' in s else ''
        items.append(f'<si><t{space}>{s_esc}</t></si>')
    body = ''.join(items)
    decl = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    return (decl +
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{len(shared)}" uniqueCount="{len(shared)}">{body}</sst>').encode('utf-8')

def set_cell_value(sheet_str, shared, cell_ref, new_value):
    idx = len(shared)
    shared.append(str(new_value))
    p1 = rf'(<c r="{re.escape(cell_ref)}"[^>]*t="s"[^>]*><v>)\d+(</v></c>)'
    s, n = re.subn(p1, rf'\g<1>{idx}\g<2>', sheet_str)
    if n: return s, shared
    p2 = rf'<c r="{re.escape(cell_ref)}"([^>]*)>.*?</c>'
    def rep(m):
        attrs = re.sub(r'\s*t="[^"]*"', '', m.group(1))
        return f'<c r="{cell_ref}"{attrs} t="s"><v>{idx}</v></c>'
    s, n = re.subn(p2, rep, sheet_str, flags=re.DOTALL)
    if n: return s, shared
    row_num = re.search(r'\d+', cell_ref).group(0)
    p3 = rf'(<row\b[^>]*\br="{row_num}"[^>]*>)(.*?)(</row>)'
    def ins(m):
        return (m.group(1) + m.group(2) +
                f'<c r="{cell_ref}" t="s"><v>{idx}</v></c>' + m.group(3))
    s, n = re.subn(p3, ins, sheet_str, flags=re.DOTALL)
    if n: return s, shared
    return sheet_str, shared

# ── ESTILOS: negro explícito en font[10] ──────────────────────────────────────
def fix_font_color_black(styles_bytes):
    styles = styles_bytes.decode('utf-8')
    fonts  = list(re.finditer(r'<font>(.*?)</font>', styles, re.DOTALL))
    if len(fonts) <= 10: return styles_bytes, False
    f10     = fonts[10]
    new_f10 = f10.group(0).replace('<color theme="1"/>', '<color rgb="FF000000"/>')
    if new_f10 == f10.group(0): return styles_bytes, False
    return (styles[:f10.start()] + new_f10 + styles[f10.end():]).encode('utf-8'), True

# ── CABECERA HRC ──────────────────────────────────────────────────────────────
def actualizar_cabecera_hrc(sheet_str, shared, datos, ciclo_col):
    celdas = {
        'C7':             datos['tramo_hrc'],       # México-Querétaro-Irapuato...
        'F7':             datos['cod_rev_general'],
        'F9':             datos['clave_ref'],
        'F10':            datos['titulo_hrc'],       # título específico HRC
        'F11':            datos['cod_documento'],
        'F12':            datos['originador'],
        'J11':            datos['revisor_ppal'],
        'J12':            datos['entidad_rev'],
        f'{ciclo_col}11': datos['fecha_ciclo'],
    }
    for cell, val in celdas.items():
        if val:
            sheet_str, shared = set_cell_value(sheet_str, shared, cell, val)
    return sheet_str, shared


# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Generador HRC", page_icon="📋", layout="centered")
st.title("📋 Generador HRC")
st.caption("Actualiza portada, headers de página y cabeceras de todos los HRC.")

st.divider()
c1, c2 = st.columns(2)
with c1:
    st.markdown("**1 · Excel de datos**")
    f_datos = st.file_uploader("datos_entrada.xlsx", type=["xlsx"], key="datos")
with c2:
    st.markdown("**2 · Plantilla HRC**")
    f_plantilla = st.file_uploader("plantilla_HRC.xlsx", type=["xlsx"], key="plantilla")

datos = None
if f_datos:
    try:
        datos = leer_datos(f_datos)
        st.divider()
        st.subheader("Datos leídos")

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Portada**")
            st.code(f"Título:  {datos['titulo_portada'] or '(vacío)'}")
            st.code(f"Tramo:   {datos['tramo_portada'] or '(vacío)'}")
            st.markdown("**Cabecera HRC**")
            st.code(f"Cód. rev:   {datos['cod_rev_general'] or '(vacío)'}")
            st.code(f"Clave ref:  {datos['clave_ref'] or '(vacío)'}")
            st.code(f"Título HRC: {datos['titulo_hrc'] or '(vacío)'}")
            st.code(f"Tramo HRC:  {datos['tramo_hrc'] or '(vacío)'}")
            st.code(f"Cód. doc:   {datos['cod_documento'] or '(vacío)'}")
        with col2:
            st.markdown("**Originador / Revisor**")
            st.code(f"Originador:  {datos['originador'] or '(vacío)'}")
            st.code(f"Revisor ppal:{datos['revisor_ppal'] or '(vacío)'}")
            st.code(f"Entidad rev: {datos['entidad_rev'] or '(vacío)'}")
            st.markdown("**Ciclo**")
            st.code(f"Revisión:    {datos['num_rev'] or '(vacío)'}")
            st.code(f"Fecha ciclo: {datos['fecha_ciclo'] or '(vacío)'}")
    except Exception as e:
        st.error(f"Error leyendo datos_entrada.xlsx: {e}")

st.divider()
with st.expander("📌 Qué se modifica", expanded=False):
    st.markdown("""
| Hoja | Elemento | Campo (datos_entrada) | Celda destino |
|------|----------|----------------------|---------------|
| PORTADA | Header No.Doc./Rev. | B20 + B2 | oddHeader |
| PORTADA | Tramo | B6 | CuadroTexto 4 run 2 |
| PORTADA | Título | B5 | CuadroTexto 4 run 3 |
| CONTROL FIRMAS | Header No.Doc./Rev. | B20 + B2 | oddHeader |
| HRC_* (todas) | Tramo | B23 | C7 |
| HRC_* (todas) | Cód. revisión gral | B20 | F7 |
| HRC_* (todas) | Clave referencia | B21 | F9 |
| HRC_* (todas) | Título HRC | B22 | F10 |
| HRC_* (todas) | Código documento | B24 | F11 |
| HRC_* (todas) | Originador | B25 | F12 |
| HRC_* (todas) | Revisor principal | B26 | J11 |
| HRC_* (todas) | Entidad Revisora | B27 | J12 |
| HRC_* (todas) | Fecha ciclo | B30 | N11/O11/P11 |
| Estilos | Font negro | — | styles.xml font[10] |
""")

generar = st.button("⚡ Generar",
                    disabled=(not f_datos or not f_plantilla or datos is None),
                    type="primary")

if generar and datos and f_plantilla:
    try:
        f_plantilla.seek(0)
        with zipfile.ZipFile(io.BytesIO(f_plantilla.read()), "r") as zin:
            infos = {item.filename: item for item in zin.infolist()}
            files = {name: zin.read(name) for name in zin.namelist()}

        sheet_map = get_sheet_map(files)
        num_rev   = datos["num_rev"].zfill(2)
        ciclo_col = {"00": "N", "01": "O"}.get(num_rev, "P")
        log = []

        # ── 1. Portada: CuadroTexto 4 ─────────────────────────────────────────
        portada = next((s for s in sheet_map if "PORTADA" in s.upper()), None)
        if portada:
            snum  = sheet_map[portada]
            dpath = get_drawing_path(files, snum)
            if dpath and dpath in files:
                new_bytes, ok, msg = actualizar_portada(
                    files[dpath], datos["titulo_portada"], datos["tramo_portada"])
                files[dpath] = new_bytes
                log.append(f"{'✅' if ok else '⚠️'} Portada drawing → {msg}")

        # ── 2. oddHeader en PORTADA y CONTROL DE FIRMAS ───────────────────────
        hojas_hdr = [s for s in sheet_map
                     if "PORTADA" in s.upper() or "CONTROL" in s.upper() or "FIRMA" in s.upper()]
        for sname in hojas_hdr:
            snum  = sheet_map[sname]
            spath = f"xl/worksheets/sheet{snum}.xml"
            if spath not in files: continue
            new_bytes, changed = actualizar_odd_header(
                files[spath], datos["cod_rev_general"], num_rev)
            files[spath] = new_bytes
            cod_base = re.sub(r"-S\d{2}$", "", datos["cod_rev_general"]).strip()
            log.append(f"{'✅' if changed else '⚠️'} Header '{sname}' → "
                       f"No. Doc. {cod_base}  Rev. S{num_rev}")

        # ── 3. Estilos: negro en font[10] ─────────────────────────────────────
        if "xl/styles.xml" in files:
            new_bytes, changed = fix_font_color_black(files["xl/styles.xml"])
            files["xl/styles.xml"] = new_bytes
            log.append(f"{'✅' if changed else 'ℹ️ ya negro'} styles.xml font[10] → FF000000")

        # ── 4. Cabeceras HRC ──────────────────────────────────────────────────
        shared    = get_shared_strings(files)
        hrc_sheets = [s for s in sheet_map if "HRC" in s.upper()]
        for sname in hrc_sheets:
            snum  = sheet_map[sname]
            spath = f"xl/worksheets/sheet{snum}.xml"
            if spath not in files: continue
            sheet_str, shared = actualizar_cabecera_hrc(
                files[spath].decode('utf-8'), shared, datos, ciclo_col)
            files[spath] = sheet_str.encode('utf-8')
            log.append(f"✅ {sname} → cabecera actualizada (ciclo {ciclo_col}, col {ciclo_col}11)")

        files["xl/sharedStrings.xml"] = rebuild_shared_strings(shared)

        # ── Reconstruir ZIP ───────────────────────────────────────────────────
        out_buf = io.BytesIO()
        with zipfile.ZipFile(out_buf, "w") as zout:
            for fname, content in files.items():
                ct = infos[fname].compress_type if fname in infos else zipfile.ZIP_DEFLATED
                zout.writestr(fname, content, compress_type=ct)
        out_buf.seek(0)

        st.success(f"✅ Completado — {len(hrc_sheets)} hojas HRC procesadas  |  Revisión S{num_rev}")
        with st.expander("Log detallado"):
            for line in log: st.text(line)

        nombre = f"{datos['cod_rev_general'] or f'HRC_S{num_rev}'}.xlsx"
        st.download_button(
            label=f"⬇️ Descargar {nombre}",
            data=out_buf.getvalue(),
            file_name=nombre,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Error: {e}")
        st.exception(e)
